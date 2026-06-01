"""Append job application records to an Excel file across runs."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from src.utils.company_tracker import normalize_company
from src.utils.formatters import excel_cell_value

SHEET_NAME = "Job Applications"

COLUMNS = [
    "Platform",
    "Run Date",
    "Run Time",
    "Job ID",
    "Job Title",
    "Company",
    "Location",
    "Experience",
    "Salary",
    "Posted Date",
    "Search Keyword",
    "Skills",
    "Job URL",
    "External Apply URL",
    "Status",
    "Applied At",
    "Notes",
]


class ExcelJobLogger:
    """Creates or appends rows to a single Excel workbook."""

    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.run_date = datetime.now().strftime("%Y-%m-%d")
        self.run_time = datetime.now().strftime("%H:%M:%S")

    def _read_sheet(self) -> tuple[dict[str, int], list] | None:
        if not self.filepath.exists():
            return None

        wb = load_workbook(self.filepath, read_only=True, data_only=True)
        try:
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            col = {h: i for i, h in enumerate(headers) if h}
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            return col, rows
        finally:
            wb.close()

    def load_applied_job_ids(self) -> set[str]:
        """Return job IDs that were successfully applied in any previous run."""
        result = self._read_sheet()
        if not result:
            return set()

        col, rows = result
        if "Job ID" not in col or "Status" not in col:
            return set()

        job_id_idx = col["Job ID"]
        status_idx = col["Status"]
        applied: set[str] = set()
        for row in rows:
            if not row or len(row) <= max(job_id_idx, status_idx):
                continue
            if row[status_idx] == "Applied" and row[job_id_idx]:
                applied.add(str(row[job_id_idx]))
        return applied

    def load_applied_companies(self) -> set[str]:
        """Return normalized company names already applied to."""
        result = self._read_sheet()
        if not result:
            return set()

        col, rows = result
        if "Company" not in col or "Status" not in col:
            return set()

        company_idx = col["Company"]
        status_idx = col["Status"]
        companies: set[str] = set()
        for row in rows:
            if not row or len(row) <= max(company_idx, status_idx):
                continue
            if row[status_idx] == "Applied" and row[company_idx]:
                norm = normalize_company(str(row[company_idx]))
                if norm:
                    companies.add(norm)
        return companies

    def is_company_applied(self, company: str, applied_companies: set[str]) -> bool:
        norm = normalize_company(company)
        return bool(norm and norm in applied_companies)

    def append_job(
        self,
        job,
        keyword: str,
        status: str,
        notes: str = "",
        applied_at: str = "",
        platform: str = "Naukri",
        external_apply_url: str = "",
    ) -> None:
        wb = self._open_or_create_workbook()
        ws = wb[SHEET_NAME]

        skills = ", ".join(job.tags) if getattr(job, "tags", None) else ""
        apply_link = getattr(job, "apply_link", "") or ""
        if not apply_link and getattr(job, "job_id", None):
            if str(job.job_id).startswith("li-"):
                apply_link = getattr(job, "job_url", "") or ""
            else:
                apply_link = f"https://www.naukri.com/job-listings-{job.job_id}"

        ws.append(
            [
                excel_cell_value(platform),
                excel_cell_value(self.run_date),
                excel_cell_value(self.run_time),
                excel_cell_value(job.job_id),
                excel_cell_value(job.title),
                excel_cell_value(job.company),
                excel_cell_value(job.location),
                excel_cell_value(getattr(job, "experience", "") or "N/A"),
                excel_cell_value(getattr(job, "salary", "") or "N/A"),
                excel_cell_value(getattr(job, "posted_date", "") or "N/A"),
                excel_cell_value(keyword),
                excel_cell_value(skills),
                excel_cell_value(apply_link),
                excel_cell_value(external_apply_url),
                excel_cell_value(status),
                excel_cell_value(applied_at),
                excel_cell_value(notes),
            ]
        )

        wb.save(self.filepath)

    def append_run_summary(self, stats: dict, total_found: int, platform: str = "") -> None:
        wb = self._open_or_create_workbook()
        ws = wb[SHEET_NAME]

        label = f"=== RUN SUMMARY ({platform}) ===" if platform else "=== RUN SUMMARY ==="
        ws.append(
            [
                platform or "",
                self.run_date,
                self.run_time,
                "",
                label,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Summary",
                "",
                (
                    f"Found={total_found}; Applied={stats.get('applied', 0)}; "
                    f"Already Applied={stats.get('skipped_applied', 0)}; "
                    f"Company Skipped={stats.get('skipped_company', 0)}; "
                    f"External={stats.get('skipped_external', 0)}; "
                    f"Failed={stats.get('failed', 0)}"
                ),
            ]
        )
        wb.save(self.filepath)

    def _open_or_create_workbook(self):
        if self.filepath.exists():
            wb = load_workbook(self.filepath)
            if SHEET_NAME not in wb.sheetnames:
                ws = wb.active
                ws.title = SHEET_NAME
            ws = wb[SHEET_NAME]
            self._ensure_headers(ws)
            return wb

        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        self._ensure_headers(ws)
        wb.save(self.filepath)
        return wb

    def _ensure_headers(self, ws) -> None:
        existing = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
        if existing == COLUMNS:
            return
        if not existing or existing[0] in (None, "Run Date", "Job ID"):
            ws.delete_rows(1, ws.max_row)
            ws.append(COLUMNS)
            for cell in ws[1]:
                cell.font = Font(bold=True)
